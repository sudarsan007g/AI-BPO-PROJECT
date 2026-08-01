import os
import base64
import tempfile
import asyncio
import datetime

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import requests
import edge_tts
from openpyxl import Workbook, load_workbook

app = Flask(__name__)
CORS(app)  # allows your GitHub Pages site to call this backend

# ---------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------
GROQ_API_KEY = os.environ.get("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")  # set this as a SECRET, never hardcode it
GROQ_URL = "https://api.groq.com/openai/v1/chat/completions"
MODEL = "llama-3.1-8b-instant"          # fast + generous free tier on Groq
VOICE = "en-US-AriaNeural"              # natural female voice (free, via edge-tts)
# Other natural female options to try: en-US-JennyNeural, en-GB-SoniaNeural, en-IN-NeerjaNeural

LOG_FILE = "calls_log.xlsx"

SYSTEM_PROMPT = (
    "You are Priya, a warm, friendly, professional customer-support voice agent. "
    "Speak the way a real human agent speaks on a phone call: natural, casual, "
    "reassuring, never robotic. Keep replies short - 1 to 3 sentences max - so the "
    "conversation feels like a real call, not an essay. Ask one question at a time."
)


def ask_groq(messages, max_tokens=150, temperature=0.6):
    if not GROQ_API_KEY:
        raise RuntimeError("GROQ_API_KEY is not set on the server")
    resp = requests.post(
        GROQ_URL,
        headers={
            "Authorization": f"Bearer {GROQ_API_KEY}",
            "Content-Type": "application/json",
        },
        json={
            "model": MODEL,
            "messages": messages,
            "temperature": temperature,
            "max_tokens": max_tokens,
        },
        timeout=30,
    )
    resp.raise_for_status()
    return resp.json()["choices"][0]["message"]["content"].strip()


async def _synthesize(text, path):
    communicate = edge_tts.Communicate(text=text, voice=VOICE)
    await communicate.save(path)


def text_to_speech_base64(text):
    with tempfile.NamedTemporaryFile(suffix=".mp3", delete=False) as tmp:
        audio_path = tmp.name
    try:
        asyncio.run(_synthesize(text, audio_path))
        with open(audio_path, "rb") as f:
            return base64.b64encode(f.read()).decode("utf-8")
    finally:
        if os.path.exists(audio_path):
            os.remove(audio_path)


# ---------------------------------------------------------------
# ROUTES
# ---------------------------------------------------------------
@app.route("/")
def home():
    return jsonify({"status": "AI BPO backend is running"})


@app.route("/speak", methods=["POST"])
def speak():
    """Turn any given text into natural speech (used for the opening greeting)."""
    data = request.get_json(force=True)
    text = data.get("text", "")
    audio_b64 = text_to_speech_base64(text)
    return jsonify({"audio_base64": audio_b64})


@app.route("/chat", methods=["POST"])
def chat():
    """Main conversation turn: user text in -> agent text + speech out."""
    data = request.get_json(force=True)
    user_message = data.get("message", "")
    history = data.get("history", [])

    history.append({"role": "user", "content": user_message})
    messages = [{"role": "system", "content": SYSTEM_PROMPT}] + history

    reply = ask_groq(messages)
    history.append({"role": "assistant", "content": reply})

    audio_b64 = text_to_speech_base64(reply)

    return jsonify({"reply": reply, "audio_base64": audio_b64, "history": history})


@app.route("/save-call", methods=["POST"])
def save_call():
    """Summarise the finished call and append one row to calls_log.xlsx."""
    data = request.get_json(force=True)
    history = data.get("history", [])
    transcript = "\n".join(f"{h['role']}: {h['content']}" for h in history)

    try:
        summary = ask_groq(
            [
                {
                    "role": "system",
                    "content": "Summarise this customer support call in one short sentence for a call log.",
                },
                {"role": "user", "content": transcript or "No conversation took place."},
            ],
            max_tokens=80,
            temperature=0.3,
        )
    except Exception:
        summary = "Summary unavailable"

    now = datetime.datetime.now()
    if os.path.exists(LOG_FILE):
        wb = load_workbook(LOG_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Calls"
        ws.append(["Date", "Time", "Summary", "Full Transcript"])

    ws.append([now.strftime("%Y-%m-%d"), now.strftime("%H:%M:%S"), summary, transcript])
    wb.save(LOG_FILE)

    return jsonify({"status": "saved", "summary": summary})


@app.route("/download-log", methods=["GET"])
def download_log():
    if not os.path.exists(LOG_FILE):
        return jsonify({"error": "No calls logged yet"}), 404
    return send_file(LOG_FILE, as_attachment=True, download_name="calls_log.xlsx")


if __name__ == "__main__":
    port = int(os.environ.get("PORT", 7860))
    app.run(host="0.0.0.0", port=port)
